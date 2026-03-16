import time
import pandas as pd
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from webdriver_manager.chrome import ChromeDriverManager


EXCEL_FILE = "warehouse_sql_question_bank.xlsx"
CHATBOT_URL = "http://localhost:8000/chatbot"


# -----------------------------
# Start Selenium (Headful Mode)
# -----------------------------
chrome_options = Options()
chrome_options.add_argument("--start-maximized")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=chrome_options
)

wait = WebDriverWait(driver, 60)

# Open login page first
driver.get("http://localhost:8000/login")

# Wait and click "Sign in as Guest"
guest_btn = wait.until(
    EC.element_to_be_clickable(
        (By.XPATH, "/html/body/div[2]/div[1]/form[1]/button[3]")
    )
)
guest_btn.click()

# wait for dashboard to load
time.sleep(5)

# Click SQL Assistant Service before asking question
sql_assistant_btn = wait.until(
    EC.element_to_be_clickable(
        (By.XPATH, "/html/body/div[1]/div[2]/button[2]")
    )
)
sql_assistant_btn.click()

time.sleep(2)

# Open chatbot page
driver.get(CHATBOT_URL)

time.sleep(5)

# -----------------------------
# Load Excel
# -----------------------------
wb = load_workbook(EXCEL_FILE)

for sheet_name in wb.sheetnames:

    if sheet_name == "Bot_Level":
        continue

    print(f"Processing sheet: {sheet_name}")

    sheet = wb[sheet_name]

    headers = [cell.value for cell in sheet[1]]

    q_col = headers.index("Question") + 1
    sql_col = headers.index("Generated SQL") + 1

    for row in range(2, sheet.max_row + 1):

        question = sheet.cell(row=row, column=q_col).value

        if not question:
            continue

        print("Question:", question)

        try:

            # -----------------------------
            # Find Input Box
            # -----------------------------
            input_box = wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "/html/body/div[1]/div[4]/div[2]/textarea")
                )
            )

            input_box.clear()
            input_box.send_keys(question)

            # -----------------------------
            # Click Send
            # -----------------------------
            send_btn = driver.find_element(
                By.XPATH,
                "/html/body/div[1]/div[4]/div[2]/button[2]"
            )

            send_btn.click()

            # -----------------------------
            # Wait for SQL block
            # -----------------------------
            sql_block = wait.until(
                EC.presence_of_all_elements_located(
                    (By.XPATH, "//pre/code")
                )
            )

            latest_sql = sql_block[-1].text

            print("Generated SQL:", latest_sql)

            # -----------------------------
            # Save SQL in Excel
            # -----------------------------
            sheet.cell(row=row, column=sql_col).value = latest_sql

            wb.save(EXCEL_FILE)

            time.sleep(3)

        except Exception as e:
            print("Error:", e)
            continue


# -----------------------------
# Finish
# -----------------------------
wb.save(EXCEL_FILE)
driver.quit()

print("All questions processed successfully.")